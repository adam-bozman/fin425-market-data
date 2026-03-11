import streamlit as st
import yfinance as yf
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import datetime

# ── Page config ───────────────────────────────────────────────
st.set_page_config(
    page_title="Market Data Downloader",
    page_icon="📈",
    layout="centered",
)

# ── Styling ───────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}

/* Page background */
.stApp {
    background-color: #0d1117;
    color: #e6edf3;
}

/* Hide default streamlit chrome */
#MainMenu, footer, header { visibility: hidden; }

/* Top banner */
.top-banner {
    background: linear-gradient(135deg, #161b22 0%, #1c2333 100%);
    border: 1px solid #30363d;
    border-radius: 12px;
    padding: 2rem 2.5rem 1.8rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.top-banner::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #2563eb, #7c3aed, #2563eb);
    background-size: 200%;
    animation: shimmer 3s linear infinite;
}
@keyframes shimmer { 0%{background-position:0%} 100%{background-position:200%} }

.banner-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.6rem;
    font-weight: 600;
    color: #f0f6fc;
    margin: 0 0 0.3rem 0;
    letter-spacing: -0.5px;
}
.banner-sub {
    font-size: 0.9rem;
    color: #8b949e;
    margin: 0;
    font-weight: 300;
}
.banner-tag {
    display: inline-block;
    background: #21262d;
    border: 1px solid #30363d;
    color: #58a6ff;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.72rem;
    padding: 2px 10px;
    border-radius: 20px;
    margin-top: 0.8rem;
}

/* Section labels */
.section-label {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.7rem;
    font-weight: 600;
    color: #58a6ff;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin: 1.6rem 0 0.6rem 0;
    padding-bottom: 0.4rem;
    border-bottom: 1px solid #21262d;
}

/* Input fields */
.stTextInput > div > div > input,
.stSelectbox > div > div > div,
.stDateInput > div > div > input {
    background-color: #161b22 !important;
    border: 1px solid #30363d !important;
    border-radius: 8px !important;
    color: #e6edf3 !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.9rem !important;
}
.stTextInput > div > div > input:focus,
.stDateInput > div > div > input:focus {
    border-color: #2563eb !important;
    box-shadow: 0 0 0 3px rgba(37,99,235,0.15) !important;
}
label, .stTextInput label, .stDateInput label, .stSelectbox label {
    color: #8b949e !important;
    font-size: 0.8rem !important;
    font-weight: 400 !important;
    font-family: 'IBM Plex Sans', sans-serif !important;
}

/* Download button */
.stDownloadButton > button {
    background: linear-gradient(135deg, #2563eb, #1d4ed8) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'IBM Plex Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
    padding: 0.7rem 2rem !important;
    width: 100% !important;
    transition: all 0.2s ease !important;
    letter-spacing: 0.3px;
}
.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #1d4ed8, #1e40af) !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 20px rgba(37,99,235,0.35) !important;
}

/* Run button */
.stButton > button {
    background: #21262d !important;
    color: #e6edf3 !important;
    border: 1px solid #30363d !important;
    border-radius: 8px !important;
    font-family: 'IBM Plex Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
    padding: 0.7rem 2rem !important;
    width: 100% !important;
    transition: all 0.2s ease !important;
}
.stButton > button:hover {
    border-color: #58a6ff !important;
    color: #58a6ff !important;
}

/* Info / success boxes */
.stAlert {
    border-radius: 8px !important;
    border-left-width: 3px !important;
}

/* Ticker badge row */
.ticker-row {
    display: flex;
    gap: 0.6rem;
    margin: 0.5rem 0 1.2rem 0;
    flex-wrap: wrap;
}
.ticker-badge {
    background: #21262d;
    border: 1px solid #30363d;
    border-radius: 6px;
    padding: 4px 12px;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.8rem;
    color: #58a6ff;
}
.ticker-badge span {
    color: #8b949e;
    font-size: 0.7rem;
    margin-left: 4px;
}

/* Preview table */
.preview-header {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.7rem;
    color: #8b949e;
    letter-spacing: 1px;
    text-transform: uppercase;
    margin: 1.5rem 0 0.5rem 0;
}
.stDataFrame { border-radius: 8px; overflow: hidden; }

/* Footer */
.footer {
    text-align: center;
    color: #484f58;
    font-size: 0.75rem;
    margin-top: 3rem;
    padding-top: 1.5rem;
    border-top: 1px solid #21262d;
    font-family: 'IBM Plex Mono', monospace;
}
</style>
""", unsafe_allow_html=True)

# ── Excel builder ─────────────────────────────────────────────
DARK_NAVY = "1F3864"
MID_BLUE  = "2E75B6"
LIGHT_BLUE = "D6E4F0"
INPUT_BG  = "EBF3FB"
STRIPE    = "F5F8FC"
WHITE     = "FFFFFF"

def thin_border(color="BDD7EE"):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def write_data_sheet(wb, sheet_name, df, ticker_label, frequency="1mo"):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    if df is None or df.empty:
        ws["A1"] = f"No data returned for {ticker_label}"
        ws["A1"].font = Font(name="Calibri", bold=True, color="C00000")
        return

    # Flatten MultiIndex columns (yfinance returns these for all intervals)
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df.columns = [str(c) for c in df.columns]

    # Standardise
    df = df.copy()
    if "Adj Close" in df.columns:
        df = df.rename(columns={"Adj Close": "Adj. Close"})
    keep = [c for c in ["Open","High","Low","Close","Adj. Close","Volume"] if c in df.columns]
    df = df[keep]
    df.index = pd.to_datetime(df.index).tz_localize(None)

    # Weekly: yfinance labels bars with the Sunday before the trading week;
    # Yahoo Finance website shows the Monday that opens the week. Shift +1 day.
    if frequency == "1wk":
        df.index = df.index + pd.DateOffset(days=1)



    col_widths = {"Date": 14, "Open": 14, "High": 14, "Low": 14,
                  "Close": 14, "Adj. Close": 14, "Volume": 16}

    # Banner
    ncols = len(keep) + 1
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    b = ws["A1"]
    b.value = f"{ticker_label}  —  Historical Prices"
    b.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    b.fill = PatternFill("solid", fgColor=DARK_NAVY)
    b.alignment = Alignment(horizontal="center", vertical="center")

    # Column headers
    ws.row_dimensions[2].height = 20
    headers = ["Date"] + keep
    for i, h in enumerate(headers):
        c = ws.cell(row=2, column=i+1, value=h)
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(i+1)].width = col_widths.get(h, 14)

    # Data rows
    price_fmt  = '#,##0.00'
    volume_fmt = '#,##0'
    date_fmt   = 'YYYY-MM-DD'

    for row_idx, (date, row) in enumerate(df.iterrows()):
        r = row_idx + 3
        ws.row_dimensions[r].height = 16
        bg = STRIPE if row_idx % 2 == 0 else WHITE

        dc = ws.cell(row=r, column=1, value=date.date())
        dc.number_format = date_fmt
        dc.font = Font(name="Calibri", size=10)
        dc.fill = PatternFill("solid", fgColor=bg)
        dc.alignment = Alignment(horizontal="center", vertical="center")
        dc.border = thin_border()

        for j, col in enumerate(keep):
            try:
                val = row[col]
                val = float(val) if pd.notna(val) else None
            except (TypeError, ValueError, KeyError):
                val = None
            c = ws.cell(row=r, column=j+2, value=val)
            c.number_format = volume_fmt if col == "Volume" else price_fmt
            c.font = Font(name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()

    ws.freeze_panes = "A3"

def write_cover_sheet(wb, firm_ticker, sp_ticker, rf_ticker,
                      start, end, frequency):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 3

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 60
    ws.row_dimensions[3].height = 8

    ws.merge_cells("B2:D2")
    t = ws["B2"]
    t.value = "Market Data Download"
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=22)
    t.fill = PatternFill("solid", fgColor=DARK_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 8
    ws.row_dimensions[5].height = 20
    ws.merge_cells("B5:D5")
    s = ws["B5"]
    s.value = "PARAMETERS"
    s.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    s.fill = PatternFill("solid", fgColor=MID_BLUE)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    freq_labels = {"1d": "Daily", "1wk": "Weekly", "1mo": "Monthly"}
    rows = [
        ("Firm Ticker",      firm_ticker,                     ""),
        ("Benchmark",        sp_ticker,                       "S&P 500"),
        ("Risk-Free Rate",   rf_ticker,                       "13-Week T-Bill yield"),
        ("Start Date",       start.strftime("%Y-%m-%d"),      ""),
        ("End Date",         end.strftime("%Y-%m-%d"),        ""),
        ("Frequency",        freq_labels.get(frequency, frequency), ""),
        ("Generated",        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"), ""),
        ("Source",           "Yahoo Finance via yfinance",    ""),
    ]
    for i, (label, value, note) in enumerate(rows):
        r = 6 + i
        ws.row_dimensions[r].height = 20
        lc = ws.cell(row=r, column=2, value=label)
        vc = ws.cell(row=r, column=3, value=value)
        nc = ws.cell(row=r, column=4, value=note)
        lc.font = Font(name="Calibri", bold=True, color="1F3864", size=10)
        vc.font = Font(name="Calibri", color="0000FF", size=10)
        nc.font = Font(name="Calibri", italic=True, color="7F7F7F", size=9)
        lc.fill = PatternFill("solid", fgColor=LIGHT_BLUE if i%2==0 else INPUT_BG)
        vc.fill = PatternFill("solid", fgColor=INPUT_BG if i%2==0 else LIGHT_BLUE)
        nc.fill = PatternFill("solid", fgColor=STRIPE)
        for c in [lc, vc, nc]:
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = thin_border()

    ws.row_dimensions[6 + len(rows)].height = 8
    r2 = 6 + len(rows) + 1
    ws.row_dimensions[r2].height = 20
    ws.merge_cells(f"B{r2}:D{r2}")
    sh = ws[f"B{r2}"]
    sh.value = "SHEETS INCLUDED"
    sh.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    sh.fill = PatternFill("solid", fgColor=MID_BLUE)
    sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    sheet_rows = [
        ("S&P 500",    sp_ticker,  "Benchmark — OHLCV + Adj. Close"),
        ("Risk-Free",  rf_ticker,  "T-Bill yield proxy"),
        (firm_ticker,  firm_ticker,"Firm — OHLCV + Adj. Close"),
    ]
    for j, (name, ticker, desc) in enumerate(sheet_rows):
        r3 = r2 + 1 + j
        ws.row_dimensions[r3].height = 19
        c1 = ws.cell(row=r3, column=2, value=name)
        c2 = ws.cell(row=r3, column=3, value=ticker)
        c3 = ws.cell(row=r3, column=4, value=desc)
        c1.font = Font(name="Calibri", bold=True, color="1F3864", size=10)
        c2.font = Font(name="Calibri", color="000000", size=10)
        c3.font = Font(name="Calibri", italic=True, color="7F7F7F", size=9)
        bg = LIGHT_BLUE if j%2==0 else STRIPE
        c1.fill = PatternFill("solid", fgColor=bg)
        c2.fill = PatternFill("solid", fgColor=INPUT_BG if j%2==0 else STRIPE)
        c3.fill = PatternFill("solid", fgColor=STRIPE)
        for c in [c1, c2, c3]:
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border = thin_border()

def build_excel(firm_ticker, sp_ticker, rf_ticker, start, end, frequency,
                df_firm, df_sp, df_rf):
    wb = openpyxl.Workbook()
    write_cover_sheet(wb, firm_ticker, sp_ticker, rf_ticker, start, end, frequency)
    write_data_sheet(wb, "S&P 500",   df_sp,   sp_ticker,   frequency)
    write_data_sheet(wb, "Risk-Free", df_rf,   rf_ticker,   frequency)
    write_data_sheet(wb, firm_ticker, df_firm, firm_ticker, frequency)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Fetch helper ──────────────────────────────────────────────
@st.cache_data(show_spinner=False, ttl=3600)
def fetch_rf(ticker, start, end, interval):
    """Fetch risk-free rate: always pull daily then resample.
    ^IRX pre-aggregated monthly/weekly data cuts off early in yfinance."""
    df = yf.download(
        ticker,
        start=start,
        end=end,
        interval="1d",
        auto_adjust=False,
        progress=False,
    )
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df.columns = [str(c) for c in df.columns]

    if df.empty:
        return df

    # Resample to target frequency using period-end close (last trading day)
    freq_map = {"1d": None, "1wk": "W-FRI", "1mo": "ME"}
    resample_rule = freq_map.get(interval)

    if resample_rule:
        agg = {
            "Open":  "first",
            "High":  "max",
            "Low":   "min",
            "Close": "last",
        }
        # Only include columns that exist
        agg = {k: v for k, v in agg.items() if k in df.columns}
        if "Adj Close" in df.columns:
            agg["Adj Close"] = "last"
        if "Volume" in df.columns:
            agg["Volume"] = "sum"
        df = df.resample(resample_rule).agg(agg).dropna(subset=["Close"])

    return df


@st.cache_data(show_spinner=False, ttl=3600)
def fetch_data(ticker, start, end, interval):
    df = yf.download(
        ticker,
        start=start,
        end=end,
        interval=interval,
        auto_adjust=False,
        progress=False,
    )
    # Always return with flattened columns so cache stores clean data
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    df.columns = [str(c) for c in df.columns]
    return df

# ── UI ────────────────────────────────────────────────────────
st.markdown("""
<div class="top-banner">
    <p class="banner-title">📈 Market Data Downloader</p>
    <p class="banner-sub">Pull historical price data from Yahoo Finance and export to a formatted Excel workbook.</p>
    <span class="banner-tag">FIN 425 · WSU Carson College of Business</span>
</div>
""", unsafe_allow_html=True)

# ── Inputs ────────────────────────────────────────────────────
st.markdown('<p class="section-label">Firm</p>', unsafe_allow_html=True)
firm_ticker = st.text_input(
    "Ticker symbol",
    value="LMT",
    placeholder="e.g. AAPL, MSFT, BA, LMT",
    label_visibility="collapsed",
).strip().upper()

st.markdown('<p class="section-label">Date Range</p>', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    start_date = st.date_input(
        "Start date",
        value=datetime.date(2020, 1, 1),
        min_value=datetime.date(1990, 1, 1),
        max_value=datetime.date.today(),
    )
with col2:
    end_date = st.date_input(
        "End date",
        value=datetime.date.today(),
        min_value=datetime.date(1990, 1, 2),
        max_value=datetime.date.today(),
    )

st.markdown('<p class="section-label">Frequency</p>', unsafe_allow_html=True)
frequency = st.selectbox(
    "Frequency",
    options=["1mo", "1wk", "1d"],
    format_func=lambda x: {"1mo": "Monthly", "1wk": "Weekly", "1d": "Daily"}[x],
    label_visibility="collapsed",
)

# Fixed tickers
SP_TICKER = "^GSPC"
RF_TICKER = "^IRX"

st.markdown(f"""
<div class="ticker-row">
    <div class="ticker-badge">{firm_ticker or "—"} <span>firm</span></div>
    <div class="ticker-badge">^GSPC <span>S&P 500</span></div>
    <div class="ticker-badge">^IRX <span>risk-free</span></div>
</div>
""", unsafe_allow_html=True)

# ── Validation ────────────────────────────────────────────────
errors = []
if not firm_ticker:
    errors.append("Please enter a firm ticker.")
if start_date >= end_date:
    errors.append("Start date must be before end date.")

for e in errors:
    st.error(e)

# ── Fetch & Download ──────────────────────────────────────────
if not errors:
    if st.button("Fetch Data & Build Excel", use_container_width=True):
        with st.spinner("Connecting to Yahoo Finance..."):
            try:
                df_firm = fetch_data(firm_ticker, start_date, end_date, frequency)
                df_sp   = fetch_data(SP_TICKER,   start_date, end_date, frequency)
                # ^IRX monthly/weekly aggregation is incomplete in yfinance;
                # fetch daily and resample ourselves for full coverage
                df_rf   = fetch_rf(RF_TICKER, start_date, end_date, frequency)

                if df_firm.empty:
                    st.error(f"No data found for **{firm_ticker}**. Check the ticker and try again.")
                    st.stop()

                # Build Excel
                excel_buf = build_excel(
                    firm_ticker, SP_TICKER, RF_TICKER,
                    start_date, end_date, frequency,
                    df_firm, df_sp, df_rf,
                )

                # Stats
                freq_label = {"1mo":"Monthly","1wk":"Weekly","1d":"Daily"}[frequency]
                n_rows = len(df_firm)

                st.success(f"✓  {n_rows} {freq_label.lower()} observations fetched for **{firm_ticker}**")

                # Preview
                st.markdown('<p class="preview-header">Preview — ' + firm_ticker + '</p>', unsafe_allow_html=True)
                preview = df_firm.copy()
                if isinstance(preview.columns, pd.MultiIndex):
                    preview.columns = preview.columns.get_level_values(0)
                preview.columns = [str(c) for c in preview.columns]
                if "Adj Close" in preview.columns:
                    preview = preview.rename(columns={"Adj Close": "Adj. Close"})
                preview.index = pd.to_datetime(preview.index).tz_localize(None)
                preview.index = preview.index.strftime("%Y-%m-%d")
                st.dataframe(
                    preview.head(8).style.format({
                        c: "{:,.2f}" for c in preview.columns if c != "Volume"
                    }).format({"Volume": "{:,.0f}"}),
                    use_container_width=True,
                )

                # Download button
                fname = f"{firm_ticker}_market_data_{start_date}_{end_date}.xlsx"
                st.download_button(
                    label="⬇  Download Excel Workbook",
                    data=excel_buf,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            except Exception as ex:
                st.error(f"Something went wrong: {ex}")

# ── Footer ────────────────────────────────────────────────────
st.markdown("""
<div class="footer">
    Data sourced from Yahoo Finance · Built with Streamlit & yfinance · FIN 425
</div>
""", unsafe_allow_html=True)
